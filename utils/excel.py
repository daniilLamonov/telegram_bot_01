from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO
from typing import List
from database.repositories import ChatRepo, OperationRepo


async def export_to_excel(
    chat_id: int = None,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
):
    import pandas as pd
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if chat_id:
        chat_info = await ChatRepo.get_chat(chat_id)
        if start_date and end_date:
            operations = await OperationRepo.get_operations_with_period(
                chat_id, start_date, end_date
            )
            chat_commissions = await OperationRepo.get_commissions_operations(chat_id, start_date, end_date)
        else:
            operations = await OperationRepo.get_operations(chat_id)
            chat_commissions = await OperationRepo.get_commissions_operations(chat_id)


        balance_data = {
            "Контрагент": [
                chat_info["contractor_name"] if chat_info else "Не установлено"
            ],
            "Комиссия %": [
                (
                    float(chat_info["commission_percent"])
                    if chat_info and chat_info["commission_percent"]
                    else 0
                )
            ],
            "Баланс RUB": [float(chat_info["balance_rub"]) if chat_info else 0],
            "Баланс USDT": [float(chat_info["balance_usdt"]) if chat_info else 0],
            "Комиссионные USDT": [chat_commissions if chat_info else 0],
            "Создан": [chat_info["created_at"] if chat_info else None],
            "Обновлен": [chat_info["updated_at"] if chat_info else None],
        }
        df_balance = pd.DataFrame(balance_data)

    else:
        all_chats = await ChatRepo.get_all_chats()

        if start_date and end_date:
            operations = await OperationRepo.get_operations_with_period(
                None, start_date, end_date
            )
        else:
            operations = await OperationRepo.get_operations()

        balance_data = []
        for chat in all_chats:
            chat_commissions = await OperationRepo.get_commissions_operations(chat["chat_id"], start_date, end_date)
            balance_data.append(
                {
                    "Chat ID": chat["chat_id"],
                    "Контрагент": chat["contractor_name"] or "Не установлено",
                    "Комиссия %": (
                        float(chat["commission_percent"])
                        if chat["commission_percent"]
                        else 0
                    ),
                    "Баланс RUB": (
                        float(chat["balance_rub"]) if chat["balance_rub"] else 0
                    ),
                    "Баланс USDT": (
                        float(chat["balance_usdt"]) if chat["balance_usdt"] else 0
                    ),
                    "Комиссионные USDT": chat_commissions if chat_commissions else 0,
                    "Создан": chat["created_at"],
                    "Обновлен": chat["updated_at"],
                }
            )

        df_balance = pd.DataFrame(balance_data)

    df_operations = (
        pd.DataFrame([dict(row) for row in operations])
        if operations
        else pd.DataFrame()
    )

    if not df_operations.empty:
        rename_dict = {
            "operation_id": "ID операции",
            "user_id": "ID пользователя",
            "username": "Пользователь",
            "operation_type": "Тип операции",
            "amount": "Сумма",
            "currency": "Валюта",
            "exchange_rate": "Курс",
            "timestamp": "Время",
            "description": "Описание",
            "contractor_name": "Контрагент",
        }
        if not chat_id:
            rename_dict["chat_id"] = "Chat ID"

        df_operations = df_operations.rename(columns=rename_dict)

        cols = df_operations.columns.tolist()
        if "Контрагент" in cols:
            cols.remove("Контрагент")
            cols = ["Контрагент"] + cols
            df_operations = df_operations[cols]

    def style_worksheet(worksheet, df, header_color="4472C4", stripe_color="D9E1F2"):

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color=header_color, end_color=header_color, fill_type="solid"
        )
        stripe_fill = PatternFill(
            start_color=stripe_color, end_color=stripe_color, fill_type="solid"
        )
        border = Border(
            left=Side(style="thin", color="B4C7E7"),
            right=Side(style="thin", color="B4C7E7"),
            top=Side(style="thin", color="B4C7E7"),
            bottom=Side(style="thin", color="B4C7E7"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical="center")

                if row_num % 2 == 0:
                    cell.fill = stripe_fill

        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)

            header_cell = worksheet.cell(row=1, column=col_num)
            max_length = len(str(header_cell.value))

            for row_num in range(2, min(102, len(df) + 2)):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.freeze_panes = "A2"

    def format_percent(percent):
        if percent == int(percent):
            return int(percent)
        return percent

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if chat_id:
            df_balance.to_excel(writer, sheet_name="Баланс чата", index=False)
            style_worksheet(
                writer.sheets["Баланс чата"],
                df_balance,
                header_color="70AD47",
                stripe_color="E2EFDA",
            )
        else:
            df_balance.to_excel(writer, sheet_name="Все чаты", index=False)
            style_worksheet(
                writer.sheets["Все чаты"],
                df_balance,
                header_color="70AD47",
                stripe_color="E2EFDA",
            )

        if not df_operations.empty:
            df_operations.to_excel(writer, sheet_name="Операции", index=False)
            worksheet = writer.sheets["Операции"]
            style_worksheet(
                worksheet, df_operations, header_color="4472C4", stripe_color="D9E1F2"
            )
            worksheet.auto_filter.ref = worksheet.dimensions

            changed_date_fill = PatternFill(
                start_color="FFCCCB",
                end_color="FFCCCB",
                fill_type="solid"
            )

            description_col_idx = None
            if "Описание" in df_operations.columns:
                description_col_idx = df_operations.columns.get_loc("Описание") + 1

                for row_num in range(2, len(df_operations) + 2):
                    description_cell = worksheet.cell(row=row_num, column=description_col_idx)
                    description_value = str(description_cell.value or "")

                    if "Дата изменена:" in description_value:
                        for col_num in range(1, len(df_operations.columns) + 1):
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.fill = changed_date_fill

            if description_col_idx:
                description_col_letter = get_column_letter(description_col_idx)

                header_cell = worksheet.cell(row=1, column=description_col_idx)
                max_length = len(str(header_cell.value))

                for row_num in range(2, len(df_operations) + 2):
                    cell = worksheet.cell(row=row_num, column=description_col_idx)
                    cell_length = len(str(cell.value or ""))
                    if cell_length > max_length:
                        max_length = cell_length

                adjusted_width = min(max_length + 2, 100)
                worksheet.column_dimensions[description_col_letter].width = adjusted_width

        # 3-й лист: Отчет
        if not df_operations.empty and "Сумма" in df_operations.columns:
            df_checks = df_operations[
                df_operations["Тип операции"] == "пополнение_руб_чек"
                ].copy()

            if not df_checks.empty:
                grouped = df_checks.groupby("Контрагент", as_index=False).agg(
                    {"Сумма": ["sum", "count"]}
                )

                grouped.columns = ["Контрагент", "Общая сумма чеков (руб)", "Количество чеков"]

                period_text = (
                    f"{start_date.strftime('%d.%m.%Y')}–{end_date.strftime('%d.%m.%Y')}"
                    if start_date
                    else "За всё время"
                )
                grouped.insert(0, "Период", period_text)

                grouped["Общая сумма чеков (руб)"] = grouped["Общая сумма чеков (руб)"].astype(float)
                grouped["Количество чеков"] = grouped["Количество чеков"].astype(int)

                total_amount = grouped["Общая сумма чеков (руб)"].sum()
                total_count = grouped["Количество чеков"].sum()

                totals_row = pd.DataFrame({
                    "Период": [period_text],
                    "Контрагент": ["ИТОГО"],
                    "Общая сумма чеков (руб)": [total_amount],
                    "Количество чеков": [total_count]
                })

                grouped = pd.concat([grouped, totals_row], ignore_index=True)

                grouped.to_excel(writer, sheet_name="Отчет", index=False)

                worksheet = writer.sheets["Отчет"]

                style_worksheet(
                    worksheet,
                    grouped.iloc[:-1],
                    header_color="C55A11",
                    stripe_color="FCE4D6",
                )

                from openpyxl.styles import Font, PatternFill, Alignment

                total_fill = PatternFill(start_color="A04000", end_color="A04000", fill_type="solid")
                total_font = Font(bold=True, color="FFFFFF")

                last_row = len(grouped) + 1
                for col in range(1, 5):
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for row in worksheet.iter_rows(min_row=2, max_row=len(grouped) + 1, min_col=3, max_col=4):
                    for cell in row:
                        if cell.row == last_row:
                            cell.number_format = '#,##0'
                        else:
                            cell.number_format = '#,##0'

        # ✅ 4-й лист: чеки для импорта (только чеки)
        if not df_operations.empty and "Сумма" in df_operations.columns:
            df_checks_import = df_operations[
                df_operations["Тип операции"] == "пополнение_руб_чек"
                ].copy()

            if not df_checks_import.empty:
                commission_map = {}
                if (
                        "Контрагент" in df_balance.columns
                        and "Комиссия %" in df_balance.columns
                ):
                    commission_map = dict(
                        zip(df_balance["Контрагент"], df_balance["Комиссия %"])
                    )

                import_data = []
                for _, row in df_checks_import.iterrows():
                    contractor = row["Контрагент"]
                    amount = float(row["Сумма"])
                    commission_percent = commission_map.get(contractor, 0.0)

                    import_data.append(
                        {
                            "A": "",
                            "B": "да",
                            "C": "",
                            "D-E": contractor,
                            "F-G": "QR",
                            "H": (int(amount) if amount == int(amount) else amount),
                            "I": f"{format_percent(commission_percent)}%",
                        }
                    )

                df_import = pd.DataFrame(import_data)

                df_import.to_excel(
                    writer,
                    sheet_name="Чеки для импорта",
                    index=False,
                    header=False,
                    startrow=0,
                )

                ws_import = writer.sheets["Чеки для импорта"]

                row_num = 1
                for idx, row_data in enumerate(import_data):
                    ws_import[f"A{row_num}"] = ""

                    ws_import[f"B{row_num}"] = "Да"
                    ws_import[f"B{row_num}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_import[f"B{row_num}"].font = Font(name='Arial', bold=True)

                    ws_import[f"C{row_num}"] = ""

                    ws_import.merge_cells(f"D{row_num}:E{row_num}")
                    ws_import[f"D{row_num}"] = row_data["D-E"]
                    ws_import[f"D{row_num}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_import[f"D{row_num}"].font = Font(name='Arial', bold=True)

                    ws_import.merge_cells(f"F{row_num}:G{row_num}")
                    ws_import[f"F{row_num}"] = "QR"
                    ws_import[f"F{row_num}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_import[f"F{row_num}"].font = Font(name='Arial', bold=True)

                    ws_import[f"H{row_num}"] = row_data["H"]
                    ws_import[f"H{row_num}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_import[f"H{row_num}"].font = Font(name='Arial')

                    commission_value = row_data["I"]
                    ws_import[f"I{row_num}"] = commission_value
                    ws_import[f"I{row_num}"].alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_import[f"I{row_num}"].font = Font(name='Arial')

                    row_num += 1

                ws_import.column_dimensions["A"].width = 5
                ws_import.column_dimensions["B"].width = 8
                ws_import.column_dimensions["C"].width = 5
                ws_import.column_dimensions["D"].width = 10
                ws_import.column_dimensions["E"].width = 10
                ws_import.column_dimensions["F"].width = 10
                ws_import.column_dimensions["G"].width = 10
                ws_import.column_dimensions["H"].width = 10
                ws_import.column_dimensions["I"].width = 10

    buffer.seek(0)
    return buffer


async def export_comparison_report(
        only_in_file: List[dict],
        only_in_db: List[dict],
) -> BytesIO:

    wb = Workbook()
    ws = wb.active
    ws.title = "Расхождения"

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")

    headers = ["Источник", "Сумма (₽)", "Дата/Время", "Контрагент/Chat", "Добавил", "ID"]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    row = 2

    if only_in_file:
        for op in only_in_file:
            ws.append([
                "Файл (нет в БД)",
                op['amount'],
                op['datetime'].strftime('%d.%m.%Y %H:%M:%S'),
                "-",
                "-",
                f"{op['transaction_id']}"
            ])
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = red_fill
            row += 1

    if only_in_db:
        for op in only_in_db:
            chat_name = await ChatRepo.get_contractor_name(op['chat_id'])
            ws.append([
                "БД (нет в файле)",
                float(op['amount']),
                op['timestamp'].strftime('%d.%m.%Y %H:%M:%S'),
                f"{chat_name}",
                op['username'],
                f"{op['operation_id']}"
            ])
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = yellow_fill
            row += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer