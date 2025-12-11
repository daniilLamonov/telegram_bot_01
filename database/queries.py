import uuid
from datetime import datetime
from typing import Optional
from .connection import get_pool


async def get_operation_details(operation_id: str) -> dict | None:
    pool = get_pool()
    async with pool.acquire() as conn:
        result = await conn.fetchrow('''
                                     SELECT *
                                     FROM operations
                                     WHERE operation_id = $1
                                     ''', operation_id)
        return dict(result) if result else None


async def delete_operation_with_balance_correction(operation_id: str, chat_id: int) -> dict:
    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            operation = await get_operation_details(operation_id)

            if not operation:
                return {
                    'success': False,
                    'message': 'Операция не найдена',
                    'operation': None
                }

            op_type = operation['operation_type']
            amount = float(operation['amount'])
            currency = operation['currency']

            balance_rub, balance_usdt = await get_balance(chat_id)

            if currency == 'RUB':
                if op_type in ['пополнение_руб', 'пополнение_руб_чек']:
                    balance_rub -= amount
                elif op_type == 'выплата_руб':
                    balance_rub += amount
                elif op_type == 'обмен_руб_usdt':
                    balance_rub += amount
                    if operation['exchange_rate']:
                        balance_usdt -= amount / float(operation['exchange_rate'])
                elif op_type == 'комиссия':
                    balance_rub += amount

            elif currency == 'USDT':
                if op_type == 'пополнение_usdt':
                    balance_usdt -= amount
                elif op_type == 'выплата_usdt':
                    balance_usdt += amount

            await update_balance(chat_id, balance_rub, balance_usdt)

            await conn.execute('''
                               DELETE
                               FROM operations
                               WHERE operation_id = $1
                               ''', operation_id)

            return {
                'success': True,
                'message': 'Операция удалена, баланс скорректирован',
                'operation': dict(operation),
                'new_balance': {'rub': balance_rub, 'usdt': balance_usdt}
            }


async def save_contractor_name(chat_id: int, contractor_name: str):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        await conn.execute('''
                           INSERT INTO chats (chat_id, contractor_name)
                           VALUES ($1, $2)
                           ON CONFLICT (chat_id) DO UPDATE SET contractor_name = $2
                           ''', chat_id, contractor_name)


async def get_contractor_name(chat_id: int) -> str:
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        result = await conn.fetchval(
            'SELECT contractor_name FROM chats WHERE chat_id = $1',
            chat_id
        )
        return result if result else "Не установлено"


async def get_check_count(chat_id: int) -> int:
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        result = await conn.fetchval('''
                                     SELECT COUNT(*)
                                     FROM operations
                                     WHERE chat_id = $1
                                       AND operation_type = 'пополнение_руб_чек'
                                     ''', chat_id)
        return int(result) if result else 0


async def get_commission(chat_id: int) -> float:
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        result = await conn.fetchval(
            'SELECT commission_percent FROM chats WHERE chat_id = $1',
            chat_id
        )
        return float(result) if result else 0.0


async def set_commission(chat_id: int, percent: float):
    db_pool = get_pool()

    async with db_pool.acquire() as conn:
        # Проверяем существует ли чат
        exists = await conn.fetchval(
            'SELECT EXISTS(SELECT 1 FROM chats WHERE chat_id = $1)',
            chat_id
        )

        if exists:
            # Если существует - только обновляем commission
            await conn.execute('''
                               UPDATE chats
                               SET commission_percent = $1
                               WHERE chat_id = $2
                               ''', percent, chat_id)
            return True
        else:
            return False

async def get_balance(chat_id: int):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        result = await conn.fetchrow('''
                                     SELECT balance_rub, balance_usdt
                                     FROM chats
                                     WHERE chat_id = $1
                                     ''', chat_id)

        if not result:
            return (0.0, 0.0)

        return (float(result['balance_rub']), float(result['balance_usdt']))


async def update_balance(chat_id: int, balance_rub: float, balance_usdt: float):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        await conn.execute('''
                           INSERT INTO chats (chat_id, balance_rub, balance_usdt, updated_at)
                           VALUES ($1, $2, $3, CURRENT_TIMESTAMP)
                           ON CONFLICT (chat_id) DO UPDATE SET balance_rub  = $2,
                                                               balance_usdt = $3,
                                                               updated_at   = CURRENT_TIMESTAMP
                           ''', chat_id, balance_rub, balance_usdt)
async def add_to_balance(chat_id: int, amount_rub: float, amount_usdt: float = 0.0):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        await conn.execute(
            '''
            INSERT INTO chats (chat_id, balance_rub, balance_usdt, updated_at)
            VALUES ($1, $2, $3, CURRENT_TIMESTAMP)
            ON CONFLICT (chat_id) DO UPDATE
            SET balance_rub  = chats.balance_rub  + EXCLUDED.balance_rub,
                balance_usdt = chats.balance_usdt + EXCLUDED.balance_usdt,
                updated_at   = CURRENT_TIMESTAMP
            ''',
            chat_id, amount_rub, amount_usdt
        )


async def log_operation(chat_id: int, user_id: int, username: str, op_type: str,
                        amount: float, currency: str, exchange_rate: Optional[float] = None,
                        description: str = '') -> str:
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        operation_id = str(uuid.uuid4())[:8]
        await conn.execute('''
                           INSERT INTO operations
                           (operation_id, chat_id, user_id, username, operation_type, amount, currency, exchange_rate,
                            description)
                           VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                           ''', operation_id, chat_id, user_id, username, op_type, amount, currency, exchange_rate,
                           description)
        return operation_id


async def get_history(chat_id: int, limit: int = 10):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        results = await conn.fetch('''
                                   SELECT operation_id,
                                          user_id,
                                          username,
                                          operation_type,
                                          amount,
                                          currency,
                                          exchange_rate,
                                          TO_CHAR(timestamp, 'YYYY-MM-DD HH24:MI:SS') as timestamp,
                                          description
                                   FROM operations
                                   WHERE chat_id = $1
                                     AND operation_type != 'пополнение_руб_чек'
                                   ORDER BY timestamp DESC
                                   LIMIT $2
                                   ''', chat_id, limit)
        return [dict(row) for row in results]


async def get_chats(chat_id: int = None):
    pool = get_pool()
    async with pool.acquire() as conn:
        if chat_id is None:
            return await conn.fetch('''
                                    SELECT chat_id,
                                           contractor_name,
                                           balance_rub,
                                           balance_usdt,
                                           commission_percent,
                                           created_at,
                                           updated_at
                                    FROM chats
                                    ORDER BY contractor_name
                                    ''')
        return await conn.fetchrow('''
                                   SELECT contractor_name,
                                          balance_rub,
                                          balance_usdt,
                                          created_at,
                                          updated_at,
                                          commission_percent
                                   FROM chats
                                   WHERE chat_id = $1
                                   ''', chat_id)


async def get_operations(chat_id: int = None):
    pool = get_pool()
    async with pool.acquire() as conn:
        if chat_id is None:
            return await conn.fetch('''
                                    SELECT o.operation_id,
                                           o.chat_id,
                                           c.contractor_name,
                                           o.user_id,
                                           o.username,
                                           o.operation_type,
                                           o.amount,
                                           o.currency,
                                           o.exchange_rate,
                                           o.timestamp,
                                           o.description
                                    FROM operations o
                                             LEFT JOIN chats c ON o.chat_id = c.chat_id
                                    ORDER BY o.timestamp DESC
                                    ''')
        return await conn.fetch('''
                                SELECT o.operation_id,
                                       o.user_id,
                                       o.username,
                                       o.operation_type,
                                       o.amount,
                                       o.currency,
                                       o.exchange_rate,
                                       o.timestamp,
                                       o.description,
                                       c.contractor_name
                                FROM operations o
                                         LEFT JOIN chats c ON o.chat_id = c.chat_id
                                WHERE o.chat_id = $1
                                ORDER BY o.timestamp DESC
                                ''', chat_id)


async def get_operations_with_period(chat_id: int,
                                     start_date: datetime,
                                     end_date: datetime):
    pool = get_pool()
    async with pool.acquire() as conn:
        if chat_id is None:
            return await conn.fetch('''
            SELECT o.operation_id,
                   o.chat_id,
                   c.contractor_name,
                   o.user_id,
                   o.username,
                   o.operation_type,
                   o.amount,
                   o.currency,
                   o.exchange_rate,
                   o.timestamp,
                   o.description
            FROM operations o
                     LEFT JOIN chats c ON o.chat_id = c.chat_id
            WHERE o.timestamp BETWEEN $1 AND $2
            ORDER BY o.timestamp DESC
                                    ''', start_date, end_date)
        return await conn.fetch('''
                                SELECT o.operation_id,
                                       o.user_id,
                                       o.username,
                                       o.operation_type,
                                       o.amount,
                                       o.currency,
                                       o.exchange_rate,
                                       o.timestamp,
                                       o.description,
                                       c.contractor_name
                                FROM operations o
                                         LEFT JOIN chats c ON o.chat_id = c.chat_id
                                WHERE o.chat_id = $1
                                  AND o.timestamp BETWEEN $2 AND $3
                                ORDER BY o.timestamp DESC
                                ''', chat_id, start_date, end_date)

async def get_check(operation_id):
    db_pool = get_pool()
    async with db_pool.acquire() as conn:
        return await conn.fetchrow('''
                                        SELECT operation_id,
                                               chat_id,
                                               username,
                                               amount,
                                               currency,
                                               timestamp,
                                               description,
                                               operation_type
                                        FROM operations
                                        WHERE operation_id = $1
                                          AND operation_type = 'пополнение_руб_чек'
                                        ''', operation_id)
async def export_to_excel(
        chat_id: int = None,
        start_date: datetime | None = None,
        end_date: datetime | None = None
):
    import pandas as pd
    from io import BytesIO
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if chat_id:
        chat_info = await get_chats(chat_id)
        if start_date and end_date:
            operations = await get_operations_with_period(chat_id, start_date, end_date)
        else:
            operations = await get_operations(chat_id)

        balance_data = {
            'Контрагент': [chat_info['contractor_name'] if chat_info else 'Не установлено'],
            'Комиссия %': [
                float(chat_info['commission_percent']) if chat_info and chat_info['commission_percent'] else 0],
            'Баланс RUB': [float(chat_info['balance_rub']) if chat_info else 0],
            'Баланс USDT': [float(chat_info['balance_usdt']) if chat_info else 0],
            'Создан': [chat_info['created_at'] if chat_info else None],
            'Обновлен': [chat_info['updated_at'] if chat_info else None]
        }
        df_balance = pd.DataFrame(balance_data)

    else:
        all_chats = await get_chats()

        if start_date and end_date:
            operations = await get_operations_with_period(None, start_date, end_date)
        else:
            operations = await get_operations()

        balance_data = []
        for chat in all_chats:
            balance_data.append({
                'Chat ID': chat['chat_id'],
                'Контрагент': chat['contractor_name'] or 'Не установлено',
                'Комиссия %': float(chat['commission_percent']) if chat['commission_percent'] else 0,
                'Баланс RUB': float(chat['balance_rub']) if chat['balance_rub'] else 0,
                'Баланс USDT': float(chat['balance_usdt']) if chat['balance_usdt'] else 0,
                'Создан': chat['created_at'],
                'Обновлен': chat['updated_at']
            })

        df_balance = pd.DataFrame(balance_data)

    # ======= DataFrame операций =======
    df_operations = pd.DataFrame([dict(row) for row in operations]) if operations else pd.DataFrame()

    if not df_operations.empty:
        rename_dict = {
            'operation_id': 'ID операции',
            'user_id': 'ID пользователя',
            'username': 'Пользователь',
            'operation_type': 'Тип операции',
            'amount': 'Сумма',
            'currency': 'Валюта',
            'exchange_rate': 'Курс',
            'timestamp': 'Время',
            'description': 'Описание',
            'contractor_name': 'Контрагент'
        }
        if not chat_id:
            rename_dict['chat_id'] = 'Chat ID'

        df_operations = df_operations.rename(columns=rename_dict)

        cols = df_operations.columns.tolist()
        if 'Контрагент' in cols:
            cols.remove('Контрагент')
            cols = ['Контрагент'] + cols
            df_operations = df_operations[cols]

    # ======= Функция для красивого оформления =======
    def style_worksheet(worksheet, df, header_color='4472C4', stripe_color='D9E1F2'):
        """Применяет красивое оформление к листу Excel"""

        # Стили
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
        stripe_fill = PatternFill(start_color=stripe_color, end_color=stripe_color, fill_type='solid')
        border = Border(
            left=Side(style='thin', color='B4C7E7'),
            right=Side(style='thin', color='B4C7E7'),
            top=Side(style='thin', color='B4C7E7'),
            bottom=Side(style='thin', color='B4C7E7')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Оформляем заголовки
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Оформляем строки данных
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical='center')

                # Чередующиеся цвета строк (зебра)
                if row_num % 2 == 0:
                    cell.fill = stripe_fill

        # Автоподбор ширины столбцов
        for col_num in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            # Проверяем длину заголовка
            header_cell = worksheet.cell(row=1, column=col_num)
            max_length = len(str(header_cell.value))

            # Проверяем длину данных (первые 100 строк для скорости)
            for row_num in range(2, min(102, len(df) + 2)):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

            # Устанавливаем ширину (с ограничением)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Закрепляем первую строку
        worksheet.freeze_panes = 'A2'

    # ✅ Функция для форматирования процента
    def format_percent(percent):
        """Если процент целый - возвращает int, иначе float"""
        if percent == int(percent):
            return int(percent)
        return percent

    # ======= Создаем Excel =======
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # 1-й лист: балансы
        if chat_id:
            df_balance.to_excel(writer, sheet_name='Баланс чата', index=False)
            style_worksheet(writer.sheets['Баланс чата'], df_balance, header_color='70AD47', stripe_color='E2EFDA')
        else:
            df_balance.to_excel(writer, sheet_name='Все чаты', index=False)
            style_worksheet(writer.sheets['Все чаты'], df_balance, header_color='70AD47', stripe_color='E2EFDA')

        # 2-й лист: операции
        if not df_operations.empty:
            df_operations.to_excel(writer, sheet_name='Операции', index=False)
            worksheet = writer.sheets['Операции']
            style_worksheet(worksheet, df_operations, header_color='4472C4', stripe_color='D9E1F2')

            # ✅ Добавляем ячейку с количеством чеков СПРАВА от таблицы
            total_checks = len(df_operations[df_operations['Тип операции'] == 'пополнение_руб_чек'])

            last_col = len(df_operations.columns)
            col_letter_label = get_column_letter(last_col + 2)  # +2 для пропуска столбца
            col_letter_value = get_column_letter(last_col + 3)

            # Записываем заголовок и значение
            worksheet[f'{col_letter_label}1'] = 'ЧЕКИ:'
            worksheet[f'{col_letter_value}1'] = total_checks

            # Стилизация для "ЧЕКИ:"
            worksheet[f'{col_letter_label}1'].font = Font(bold=True, size=12, color='FFFFFF')
            worksheet[f'{col_letter_label}1'].fill = PatternFill(start_color='FF6B35', end_color='FF6B35',
                                                                 fill_type='solid')
            worksheet[f'{col_letter_label}1'].alignment = Alignment(horizontal='right', vertical='center')
            worksheet[f'{col_letter_label}1'].border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )

            # Стилизация для значения
            worksheet[f'{col_letter_value}1'].font = Font(bold=True, size=14, color='FFFFFF')
            worksheet[f'{col_letter_value}1'].fill = PatternFill(start_color='FF6B35', end_color='FF6B35',
                                                                 fill_type='solid')
            worksheet[f'{col_letter_value}1'].alignment = Alignment(horizontal='center', vertical='center')
            worksheet[f'{col_letter_value}1'].border = Border(
                left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick')
            )

            # Ширина столбцов
            worksheet.column_dimensions[col_letter_label].width = 10
            worksheet.column_dimensions[col_letter_value].width = 8

        # 3-й лист: выдача
        if not df_operations.empty and 'Сумма' in df_operations.columns:
            df_checks = df_operations[df_operations['Тип операции'] == 'пополнение_руб_чек'].copy()

            if not df_checks.empty:
                commission_map = {}
                if 'Контрагент' in df_balance.columns and 'Комиссия %' in df_balance.columns:
                    commission_map = dict(zip(df_balance['Контрагент'], df_balance['Комиссия %']))

                df_checks['Комиссия %'] = df_checks['Контрагент'].map(commission_map).fillna(0.0)

                grouped = df_checks.groupby('Контрагент', as_index=False).agg({
                    'Сумма': 'sum',
                    'Комиссия %': 'max'
                })

                grouped['Сумма'] = grouped['Сумма'].astype(float)
                grouped['Комиссия %'] = grouped['Комиссия %'].astype(float)

                grouped['Сумма комиссии'] = grouped['Сумма'] * grouped['Комиссия %'] / 100
                grouped['К выдаче'] = grouped['Сумма'] - grouped['Сумма комиссии']

                period_text = (
                    f"{start_date.strftime('%d.%m.%Y')}–{end_date.strftime('%d.%m.%Y')}"
                    if start_date else "За всё время"
                )
                grouped.insert(0, 'Период', period_text)

                grouped.rename(columns={
                    'Сумма': 'Общая сумма чеков (руб)',
                    'Комиссия %': 'Комиссия (%)',
                    'Сумма комиссии': 'Сумма комиссии (руб)',
                    'К выдаче': 'К выдаче (руб)'
                }, inplace=True)

                grouped.to_excel(writer, sheet_name='Выдача', index=False)
                style_worksheet(writer.sheets['Выдача'], grouped, header_color='C55A11', stripe_color='FCE4D6')

        # ✅ 4-й лист: чеки для импорта (только чеки)
        if not df_operations.empty and 'Сумма' in df_operations.columns:
            df_checks_import = df_operations[df_operations['Тип операции'] == 'пополнение_руб_чек'].copy()

            if not df_checks_import.empty:
                # Получаем комиссии для каждого контрагента
                commission_map = {}
                if 'Контрагент' in df_balance.columns and 'Комиссия %' in df_balance.columns:
                    commission_map = dict(zip(df_balance['Контрагент'], df_balance['Комиссия %']))

                # Формируем данные для 4-го листа
                import_data = []
                for _, row in df_checks_import.iterrows():
                    contractor = row['Контрагент']
                    amount = float(row['Сумма'])
                    commission_percent = commission_map.get(contractor, 0.0)

                    import_data.append({
                        'A': '',  # Пусто
                        'B': 'да',  # да
                        'C': '',  # Пусто
                        'D-E': contractor,  # Контрагент (объединённые столбцы D и E)
                        'F-G': 'QR',  # QR (объединённые столбцы F и G)
                        'H': amount,  # Сумма чека
                        'I': f'{format_percent(commission_percent)}%'  # ✅ Процент (целый без точки)
                    })

                df_import = pd.DataFrame(import_data)

                # Записываем БЕЗ заголовков
                df_import.to_excel(writer, sheet_name='Чеки для импорта', index=False, header=False, startrow=0)

                # Ручное оформление
                ws_import = writer.sheets['Чеки для импорта']

                # Объединяем ячейки и заполняем данные
                row_num = 1
                for idx, row_data in enumerate(import_data):
                    # A - пусто
                    ws_import[f'A{row_num}'] = ''

                    # B - "да"
                    ws_import[f'B{row_num}'] = 'Да'
                    ws_import[f'B{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_import[f'B{row_num}'].font = Font(bold=True)

                    # C - пусто
                    ws_import[f'C{row_num}'] = ''

                    # D-E - контрагент (объединяем)
                    ws_import.merge_cells(f'D{row_num}:E{row_num}')
                    ws_import[f'D{row_num}'] = row_data['D-E']
                    ws_import[f'D{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_import[f'D{row_num}'].font = Font(bold=True)

                    # F-G - QR (объединяем)
                    ws_import.merge_cells(f'F{row_num}:G{row_num}')
                    ws_import[f'F{row_num}'] = 'QR'
                    ws_import[f'F{row_num}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_import[f'F{row_num}'].font = Font(bold=True)

                    # H - сумма чека
                    ws_import[f'H{row_num}'] = row_data['H']
                    ws_import[f'H{row_num}'].alignment = Alignment(horizontal='right', vertical='center')
                    ws_import[f'H{row_num}'].number_format = '0.00'

                    # I - процент комиссии ✅
                    commission_value = row_data['I']
                    ws_import[f'I{row_num}'] = commission_value
                    ws_import[f'I{row_num}'].alignment = Alignment(horizontal='right', vertical='center')

                    # ✅ Формат: если целое - без точки, если дробное - с точкой
                    if isinstance(commission_value, int):
                        ws_import[f'I{row_num}'].number_format = '0'
                    else:
                        ws_import[f'I{row_num}'].number_format = '0.00'

                    row_num += 1

                # Ширина столбцов
                ws_import.column_dimensions['A'].width = 5
                ws_import.column_dimensions['B'].width = 8
                ws_import.column_dimensions['C'].width = 5
                ws_import.column_dimensions['D'].width = 10
                ws_import.column_dimensions['E'].width = 10
                ws_import.column_dimensions['F'].width = 10
                ws_import.column_dimensions['G'].width = 10
                ws_import.column_dimensions['H'].width = 10
                ws_import.column_dimensions['I'].width = 10

    buffer.seek(0)
    return buffer


async def get_checks_by_date(chat_id: int, start_date, end_date):
    """Получить все чеки за указанный период"""
    pool = get_pool()
    async with pool.acquire() as conn:
        results = await conn.fetch('''
            SELECT operation_id,
                   username,
                   amount,
                   timestamp,
                   description
            FROM operations
            WHERE chat_id = $1
              AND operation_type = 'пополнение_руб_чек'
              AND timestamp >= $2
              AND timestamp < $3
            ORDER BY timestamp DESC
        ''', chat_id, start_date, end_date)

        return [dict(row) for row in results]


async def get_checks_by_exact_date(chat_id: int, date):
    """Получить все чеки за конкретную дату"""
    from datetime import timedelta

    start_date = date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = start_date + timedelta(days=1)

    return await get_checks_by_date(chat_id, start_date, end_date)


async def is_chat_initialized(chat_id: int) -> bool:
    pool = get_pool()
    async with pool.acquire() as conn:
        result = await conn.fetchval(
            'SELECT EXISTS(SELECT 1 FROM chats WHERE chat_id = $1 AND is_active = true)',
            chat_id
        )
        return result


async def get_chat_info(chat_id: int) -> dict | None:
    pool = get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            'SELECT * FROM chats WHERE chat_id = $1',
            chat_id
        )
        return dict(row) if row else None


async def initialize_chat(
    chat_id: int,
    chat_title: str,
    chat_type: str,
    contractor_name: str,
    initialized_by: int
) -> bool:
    """Инициализация чата"""
    pool = get_pool()
    async with pool.acquire() as conn:
        try:
            await conn.execute('''
                INSERT INTO chats (
                    chat_id, 
                    chat_title, 
                    chat_type, 
                    contractor_name,
                    initialized_by,
                    is_active
                )
                VALUES ($1, $2, $3, $4, $5, true)
                ON CONFLICT (chat_id) 
                DO UPDATE SET 
                    contractor_name = EXCLUDED.contractor_name,
                    chat_title = EXCLUDED.chat_title,
                    is_active = true,
                    updated_at = NOW()
            ''', chat_id, chat_title, chat_type, contractor_name, initialized_by)
            return True
        except Exception as e:
            print(f"❌ Error initializing chat: {e}")
            return False
